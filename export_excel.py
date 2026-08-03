#!/usr/bin/env python3
"""
export_excel.py  --  Refresh the ONE central Excel sheet from songs.db

Double-click this file any time you want the sheet brought up to date.
It reads songs.db sitting next to it and OVERWRITES a single file,
song_review.xlsx, so there is always exactly one current sheet -- no
pile of dated snapshots.

Safe to re-run: if you've already typed answers into "Keep or Skip?" for
any row, this script reads the existing sheet first and carries those
answers forward before overwriting, so your typed decisions are never
lost by re-running it.

Columns match the approved sheet exactly:
  #, Title, Artist, Status, Instagram length, YouTube length, Length gap,
  Why flagged / notes, YouTube link, Keep or Skip?, Date added

No internet, no API keys. Needs Python + the 'openpyxl' add-on for the
colour version; falls back to a plain .csv (song_review.csv) if that
add-on isn't installed.
"""

import sqlite3
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
DB_PATH = HERE / "songs.db"
XLSX_OUT = HERE / "song_review.xlsx"
CSV_OUT = HERE / "song_review.csv"

COLUMNS = ["#", "Title", "Artist", "Status", "Instagram length",
           "YouTube length", "Length gap", "Why flagged / notes",
           "YouTube link", "Keep or Skip?", "Date added"]

# Known review-song details (found length / gap), same as the original
# hand-built sheet. Anything not listed here (new songs added later) just
# shows blank for these two columns until reviewed.
KNOWN_REVIEW_DETAILS = {
    "Birds":     {"yt_len": "4:40", "gap": "31s longer"},
    "Stratus":   {"yt_len": "3:57", "gap": "11s longer"},
    "Haru":      {"yt_len": "3:15", "gap": "10s longer"},
    "of bliss":  {"yt_len": "2:32", "gap": "38s longer"},
}


def load_existing_keep_skip():
    """Read the current song_review.xlsx (or .csv), if present, and return
    a dict of {(title, artist): 'Keep' or 'Skip' or whatever text the user
    typed} so re-running the export never erases a decision the user
    already made but hasn't synced back to the database yet."""
    carried = {}
    if XLSX_OUT.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(XLSX_OUT)
            ws = wb.active
            hrow = 5
            headers = [ws.cell(row=hrow, column=c).value for c in range(1, ws.max_column + 1)]
            if "Title" in headers and "Artist" in headers and "Keep or Skip?" in headers:
                ti = headers.index("Title") + 1
                ai = headers.index("Artist") + 1
                ki = headers.index("Keep or Skip?") + 1
                for r in range(hrow + 1, ws.max_row + 1):
                    title = ws.cell(row=r, column=ti).value
                    artist = ws.cell(row=r, column=ai).value
                    keep = ws.cell(row=r, column=ki).value
                    if title and keep:
                        carried[(title, artist)] = keep
        except Exception:
            pass  # if anything about the old file is unreadable, just skip carrying forward
    elif CSV_OUT.exists():
        try:
            import csv
            with open(CSV_OUT, encoding="utf-8-sig") as f:
                for row in csv.DictReader(f):
                    if row.get("Title") and row.get("Keep or Skip?"):
                        carried[(row["Title"], row.get("Artist"))] = row["Keep or Skip?"]
        except Exception:
            pass
    return carried


def load_rows():
    if not DB_PATH.exists():
        print(f"\n  ERROR: I can't find songs.db in this folder:\n    {HERE}")
        print("  Make sure export_excel.py and songs.db are in the SAME folder.\n")
        return None
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = list(conn.execute("SELECT * FROM songs ORDER BY id"))
    conn.close()
    return rows


def status_label(status):
    return {
        "added": "Added",
        "needs_review": "NEEDS REVIEW",
        "to_add": "To add",
        "pending": "Pending",
        "skipped": "Skipped",
    }.get(status, status)


def row_values(r, carried=None, display_num=None):
    carried = carried or {}
    is_review = r["match_status"] == "needs_review"
    known = KNOWN_REVIEW_DETAILS.get(r["title"], {})
    yt_len = (known.get("yt_len") or "") if is_review else ("matched" if r["match_status"] == "added" else "")
    gap = (known.get("gap") or "") if is_review else ""
    why = r["notes"] or ("Official channel + length matched \u2013 added automatically."
                          if r["match_status"] == "added" else "")
    if is_review:
        # carry forward whatever the user already typed, if anything
        keep_skip = carried.get((r["title"], r["artist"]), "")
    else:
        keep_skip = "Added" if r["match_status"] == "added" else ""
    return [
        display_num if display_num is not None else r["id"], r["title"], r["artist"], status_label(r["match_status"]),
        r["duration"] or "", yt_len, gap, why,
        r["youtube_url"] or "", keep_skip, r["date_captured"] or "",
    ]


def write_xlsx(rows, carried):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    out = XLSX_OUT
    wb = Workbook()
    ws = wb.active
    ws.title = "Song Sync"

    ARIAL = "Arial"
    navy = "1F3864"
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    input_fill = PatternFill("solid", fgColor="FCE4D6")
    header_fill = PatternFill("solid", fgColor=navy)
    added_fill = PatternFill("solid", fgColor="E2EFDA")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_review = sum(1 for r in rows if r["match_status"] == "needs_review")
    n_added = sum(1 for r in rows if r["match_status"] == "added")

    ws["A1"] = "Instagram \u2192 YouTube  \u2014  Song Sync Status"
    ws["A1"].font = Font(name=ARIAL, size=15, bold=True, color=navy)
    ws["A2"] = f"Mirrors songs.db as of last refresh \u00b7 {n_added} added \u00b7 {n_review} need review"
    ws["A2"].font = Font(name=ARIAL, size=10, italic=True, color="595959")
    ws["A3"] = ("How to use the amber rows: open the link, listen ~30 seconds, then type Keep or Skip "
                "in that column. Green rows are already done \u2014 nothing to do there.")
    ws["A3"].font = Font(name=ARIAL, size=10, color="595959")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells("A3:K3")
    ws.row_dimensions[3].height = 30

    hrow = 5
    for c, h in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=hrow, column=c, value=h)
        cell.font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[hrow].height = 28

    r = hrow + 1
    for i, row in enumerate(rows, start=1):
        is_review = row["match_status"] == "needs_review"
        values = row_values(row, carried, display_num=i)
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name=ARIAL, size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(c == 8))
            if is_review:
                cell.fill = review_fill
        scell = ws.cell(row=r, column=4)
        if is_review:
            scell.font = Font(name=ARIAL, size=10, bold=True, color="BF8F00")
        elif row["match_status"] == "added":
            scell.font = Font(name=ARIAL, size=10, bold=True, color="548235")
            scell.fill = added_fill
        lcell = ws.cell(row=r, column=9)
        if lcell.value:
            lcell.hyperlink = lcell.value
            lcell.font = Font(name=ARIAL, size=10, color="0563C1", underline="single")
        kcell = ws.cell(row=r, column=10)
        if is_review:
            kcell.fill = input_fill
            kcell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 42 if is_review else 22
        r += 1

    widths = [4, 24, 20, 15, 16, 15, 13, 46, 44, 14, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A{hrow}:K{r-1}"
    wb.save(out)
    return out


def write_csv(rows, carried):
    import csv
    out = CSV_OUT
    with open(out, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(COLUMNS)
        for i, row in enumerate(rows, start=1):
            w.writerow(row_values(row, carried, display_num=i))
    return out


def summarise(rows):
    from collections import Counter
    counts = Counter(r["match_status"] for r in rows)
    order = ["added", "to_add", "needs_review", "pending", "skipped"]
    return ", ".join(f"{counts[s]} {s}" for s in order if counts.get(s))


def main():
    rows = load_rows()
    if rows is None:
        return 1
    carried = load_existing_keep_skip()
    try:
        out = write_xlsx(rows, carried)
        kind = "Excel (colour-coded)"
        extra = ""
        # xlsx and csv can't both linger with stale data -- remove old csv if present
        if CSV_OUT.exists():
            CSV_OUT.unlink()
    except ImportError:
        out = write_csv(rows, carried)
        kind = "CSV (plain \u2014 opens in Excel)"
        extra = ("\n  Tip: for the colour-coded Excel version, install the add-on once by\n"
                 "  opening Command Prompt and running:   pip install openpyxl\n")

    print("\n  Done! song_review is up to date:")
    print(f"    File:   {out.name}")
    print(f"    Format: {kind}")
    print(f"    Songs:  {len(rows)}  ({summarise(rows)})")
    if carried:
        print(f"    Carried forward {len(carried)} existing Keep/Skip answer(s) from the previous version.")
    print(extra)
    print("  (This file was overwritten in place \u2014 it always mirrors the current database.)\n")
    return 0


if __name__ == "__main__":
    code = main()
    try:
        input("  Press Enter to close this window...")
    except EOFError:
        pass
    sys.exit(code)
